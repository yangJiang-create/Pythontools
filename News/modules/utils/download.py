# encoding: utf-8
import openpyxl
import os

class download():
    def download(self,key,values):
        self.path = os.path.join(os.path.abspath(__file__).split('.')[0][:-8],
                            'download.xlsx')

        if not os.path.exists(self.path):
            wb = openpyxl.Workbook(self.path)
        else:
            wb = openpyxl.load_workbook(self.path)
            if key in  wb.sheetnames :
                del wb[key]
                
        ws = wb.create_sheet(key)
        for value in values:
            ws.append(list(value))
        wb.save(self.path)
    
    def getdownload(self,mail):
        values = mail[1]
        data = []

        self.path = os.path.join(os.path.abspath(__file__).split('.')[0][:-8],
                            'download.xlsx')
        wb = openpyxl.load_workbook(self.path)
        for key in mail[2]:
            ws = wb[key]
            for row in ws.iter_rows(max_row=values,values_only=True):
                data.append(row)
        return data